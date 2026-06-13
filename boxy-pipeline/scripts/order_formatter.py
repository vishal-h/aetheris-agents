#!/usr/bin/env python3
"""Write a populated Boxy Order Form Excel file from a list of ResolvedItem dicts.

Reads a JSON array of ResolvedItem dicts from stdin (t2 output).
Writes {output_dir}/{project_name}_order_form.xlsx.
"""
import argparse
import json
import sys
from pathlib import Path

import openpyxl

SHEET_NAME = "2000 Order Form"
FIRST_DATA_ROW = 12

# Column positions (1-indexed, matching the Order Form layout)
COL_LINE = 1    # A — line number
COL_ITEM = 2    # B — *ITEM
COL_COLOR = 3   # C — *Color
COL_QTY = 4     # D — *QTY
COL_PRICE = 5   # E — Unit Price
COL_AMOUNT = 6  # F — Amount (formula)
COL_SPECIAL = 11  # K — Modification or Special Request Details

FEE_PLACEHOLDERS = ("Assembly Fee", "Modification Fee", "Delivery Fee")


def _amount_formula(row: int) -> str:
    return f'=IFERROR($D{row}*$E{row},"")'


def write_order_form(
    items: list[dict],
    template_path: Path,
    project_name: str,
    output_dir: Path,
) -> Path:
    """Fill the Order Form template and save to output_dir.

    Returns the path of the written file.
    """
    wb = openpyxl.load_workbook(template_path)
    for sheet_name in [s for s in wb.sheetnames if s != SHEET_NAME]:
        del wb[sheet_name]
    ws = wb[SHEET_NAME]

    row = FIRST_DATA_ROW
    line_num = 1

    for item in items:
        comp = item["component"]
        cat = item.get("catalog_item")
        confidence = item.get("match_confidence", "unresolved")

        ws.cell(row, COL_LINE).value = line_num
        ws.cell(row, COL_ITEM).value = comp["code"]
        ws.cell(row, COL_QTY).value = item["qty"]

        if cat is not None:
            ws.cell(row, COL_COLOR).value = cat["color_code"]
            ws.cell(row, COL_PRICE).value = item["unit_price"]
        else:
            ws.cell(row, COL_COLOR).value = None
            ws.cell(row, COL_PRICE).value = None

        ws.cell(row, COL_AMOUNT).value = _amount_formula(row)

        notes: list[str] = []
        if item.get("match_notes"):
            notes.append(item["match_notes"])
        if confidence == "unresolved":
            notes.append("UNRESOLVED - manual review required")
        ws.cell(row, COL_SPECIAL).value = "; ".join(notes) if notes else None

        row += 1
        line_num += 1

    # Fee placeholder rows — ITEM only, all other columns cleared
    for fee_name in FEE_PLACEHOLDERS:
        ws.cell(row, COL_LINE).value = line_num
        ws.cell(row, COL_ITEM).value = fee_name
        for col in (COL_COLOR, COL_QTY, COL_PRICE, COL_AMOUNT, COL_SPECIAL):
            ws.cell(row, col).value = None
        row += 1
        line_num += 1

    # Clear stale template formulas in rows beyond what we wrote
    for r in range(row, 68):
        for c in range(COL_ITEM, COL_SPECIAL + 1):
            ws.cell(r, c).value = None

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"{project_name}_order_form.xlsx"
    wb.save(out_path)
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Write Boxy Order Form from resolved items.")
    parser.add_argument("--template", required=True, type=Path, metavar="XLSX")
    parser.add_argument("--project", required=True, metavar="NAME")
    parser.add_argument("--output-dir", default="output", type=Path, metavar="DIR")
    args = parser.parse_args()

    if not args.template.exists():
        print(f"Error: template not found: {args.template}", file=sys.stderr)
        sys.exit(1)

    raw = sys.stdin.read()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        print(f"Error: invalid JSON on stdin: {exc}", file=sys.stderr)
        sys.exit(1)

    if not isinstance(data, list):
        print("Error: expected a JSON array of ResolvedItem dicts on stdin", file=sys.stderr)
        sys.exit(1)

    try:
        out_path = write_order_form(data, args.template, args.project, args.output_dir)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)

    print(out_path)


if __name__ == "__main__":
    main()
