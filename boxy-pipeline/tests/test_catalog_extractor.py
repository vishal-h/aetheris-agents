"""Tests for scripts/catalog_extractor.py — boxy-pipeline-1a t1."""
import json
import subprocess
import sys
from dataclasses import asdict
from pathlib import Path

import openpyxl
import pytest

from catalog_extractor import extract_catalog
from schema import CatalogEntry

USE_CASE_ROOT = Path(__file__).parent.parent
SAMPLES_DIR = USE_CASE_ROOT / "data" / "samples"
CATALOG_FILE = SAMPLES_DIR / "Updated_Boxy_MSRP_Sales_Order_Form.xlsx"


# ---------------------------------------------------------------------------
# Minimal Excel fixture — unit tests without real sample
# ---------------------------------------------------------------------------

@pytest.fixture()
def minimal_catalog_xlsx(tmp_path: Path) -> Path:
    """Minimal two-item Price List for unit testing extract_catalog()."""
    wb = openpyxl.Workbook()

    # 2000 Price List sheet
    ws = wb.active
    ws.title = "2000 Price List"
    # Row 1 (pandas index 0): padding
    # Row 2 (pandas index 1): color headers at pandas cols 6 and 7 (openpyxl cols 7 and 8)
    ws.cell(2, 7).value = "2001\nIvory White\nMSRP"
    ws.cell(2, 8).value = "2004\nMingo Oak\nMSRP"
    # Row 3 (pandas index 2): item W2739 — cols E=5, F=6 (openpyxl), prices at G=7, H=8
    ws.cell(3, 5).value = "W2739"
    ws.cell(3, 6).value = 'Wall Cabinet, 27"W x 39"H x 12"D, 2 Doors'
    ws.cell(3, 7).value = 450.8   # color 2001
    ws.cell(3, 8).value = 480.0   # color 2004
    # Row 4: item 3DB30 (leading digit should be stripped to base_code DB30)
    ws.cell(4, 5).value = "3DB30"
    ws.cell(4, 6).value = 'Base Cabinet, 30"W x 34"H x 24"D, 3 Drawers'
    ws.cell(4, 7).value = 1100.0  # color 2001
    ws.cell(4, 8).value = 1188.4  # color 2004

    # 3000 Price List sheet (minimal — one item)
    ws3 = wb.create_sheet("3000 Price List")
    ws3.cell(2, 7).value = "2001\nIvory White\nMSRP"
    ws3.cell(3, 5).value = "W3039"
    ws3.cell(3, 6).value = 'Wall Cabinet, 30"W x 39"H x 12"D'
    ws3.cell(3, 7).value = 520.0

    path = tmp_path / "minimal_catalog.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Unit tests (minimal fixture — no real sample needed)
# ---------------------------------------------------------------------------


def test_extract_returns_catalog_entries(minimal_catalog_xlsx):
    entries = extract_catalog(minimal_catalog_xlsx)
    assert all(isinstance(e, CatalogEntry) for e in entries)


def test_entry_count_matches_items_times_colors(minimal_catalog_xlsx):
    entries = extract_catalog(minimal_catalog_xlsx)
    # 2 items × 2 colors + 1 item × 1 color = 5
    assert len(entries) == 5


def test_sku_format(minimal_catalog_xlsx):
    entries = extract_catalog(minimal_catalog_xlsx)
    for e in entries:
        assert "-" in e.sku
        parts = e.sku.rsplit("-", 1)
        assert parts[0] == e.raw_code
        assert parts[1] == e.color_code


def test_base_code_strips_leading_digit(minimal_catalog_xlsx):
    entries = extract_catalog(minimal_catalog_xlsx)
    db30 = [e for e in entries if e.raw_code == "3DB30"]
    assert db30, "3DB30 not found"
    assert all(e.base_code == "DB30" for e in db30)


def test_base_code_unchanged_when_no_leading_digit(minimal_catalog_xlsx):
    entries = extract_catalog(minimal_catalog_xlsx)
    w2739 = [e for e in entries if e.raw_code == "W2739"]
    assert w2739
    assert all(e.base_code == "W2739" for e in w2739)


def test_enrichment_fields_empty_at_extraction(minimal_catalog_xlsx):
    entries = extract_catalog(minimal_catalog_xlsx)
    for e in entries:
        assert e.mapped_20_20_codes == []
        assert e.notes is None


def test_catalog_version_is_iso_date(minimal_catalog_xlsx):
    import datetime
    entries = extract_catalog(minimal_catalog_xlsx)
    for e in entries:
        datetime.date.fromisoformat(e.catalog_version)  # raises if invalid


def test_source_file_is_basename(minimal_catalog_xlsx):
    entries = extract_catalog(minimal_catalog_xlsx)
    for e in entries:
        assert e.source_file == minimal_catalog_xlsx.name
        assert "/" not in e.source_file


def test_series_extracted_correctly(minimal_catalog_xlsx):
    entries = extract_catalog(minimal_catalog_xlsx)
    series_found = {e.series for e in entries}
    assert "2000" in series_found
    assert "3000" in series_found


def test_dimensions_parsed(minimal_catalog_xlsx):
    entries = extract_catalog(minimal_catalog_xlsx)
    w2739 = next(e for e in entries if e.raw_code == "W2739" and e.color_code == "2001")
    assert w2739.width_in == pytest.approx(27.0)
    assert w2739.height_in == pytest.approx(39.0)
    assert w2739.depth_in == pytest.approx(12.0)


def test_cabinet_type_extracted(minimal_catalog_xlsx):
    entries = extract_catalog(minimal_catalog_xlsx)
    w2739 = next(e for e in entries if e.raw_code == "W2739" and e.color_code == "2001")
    assert w2739.cabinet_type == "Wall Cabinet"


def test_color_name_extracted(minimal_catalog_xlsx):
    entries = extract_catalog(minimal_catalog_xlsx)
    w2739_2001 = next(e for e in entries if e.raw_code == "W2739" and e.color_code == "2001")
    assert w2739_2001.color_name == "Ivory White"


def test_msrp_value(minimal_catalog_xlsx):
    entries = extract_catalog(minimal_catalog_xlsx)
    w2739_2001 = next(e for e in entries if e.raw_code == "W2739" and e.color_code == "2001")
    assert w2739_2001.msrp == pytest.approx(450.8)


def test_entry_serializes_to_jsonl(minimal_catalog_xlsx):
    entries = extract_catalog(minimal_catalog_xlsx)
    for e in entries:
        line = json.dumps(asdict(e))
        parsed = json.loads(line)
        assert parsed["sku"] == e.sku
        assert parsed["mapped_20_20_codes"] == []
        assert parsed["notes"] is None


# ---------------------------------------------------------------------------
# Integration tests — require real sample files
# ---------------------------------------------------------------------------


@pytest.mark.integration
def test_all_five_series_present():
    entries = extract_catalog(CATALOG_FILE)
    series_found = {e.series for e in entries}
    for expected in ("1000", "2000", "3000", "4000", "5000"):
        assert expected in series_found, f"Series {expected} missing from catalog"


@pytest.mark.integration
def test_total_entry_count():
    entries = extract_catalog(CATALOG_FILE)
    assert len(entries) > 200, f"Expected >200 entries, got {len(entries)}"


@pytest.mark.integration
def test_w2739_spot_check():
    entries = extract_catalog(CATALOG_FILE)
    w2739 = [e for e in entries if e.sku == "W2739-2001"]
    assert w2739, "W2739-2001 missing from catalog"
    assert w2739[0].msrp == pytest.approx(450.8), f"Wrong MSRP: {w2739[0].msrp}"
    assert w2739[0].base_code == "W2739"
    assert w2739[0].cabinet_type == "Wall Cabinet"
    assert w2739[0].series == "2000"


@pytest.mark.integration
def test_enrichment_fields_empty_real_catalog():
    entries = extract_catalog(CATALOG_FILE)
    bad = [e for e in entries if e.mapped_20_20_codes != [] or e.notes is not None]
    assert not bad, f"{len(bad)} entries have non-empty enrichment fields"


@pytest.mark.integration
def test_all_entries_have_catalog_version():
    import datetime
    entries = extract_catalog(CATALOG_FILE)
    for e in entries:
        datetime.date.fromisoformat(e.catalog_version)


@pytest.mark.integration
def test_cli_creates_jsonl(tmp_path):
    out = tmp_path / "catalog.jsonl"
    result = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "catalog_extractor.py"),
         "--catalog", str(CATALOG_FILE),
         "--output", str(out)],
        capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    assert result.returncode == 0, f"CLI failed:\n{result.stderr}"
    assert out.exists()
    lines = out.read_text().strip().splitlines()
    assert len(lines) > 200
    # Each line must be valid JSON with required fields
    first = json.loads(lines[0])
    for field in ("sku", "base_code", "raw_code", "series", "color_code",
                  "color_name", "description", "cabinet_type", "msrp",
                  "mapped_20_20_codes", "notes", "catalog_version", "source_file"):
        assert field in first, f"Missing field: {field}"


@pytest.mark.integration
def test_cli_summary_output():
    result = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "catalog_extractor.py"),
         "--catalog", str(CATALOG_FILE),
         "--output", "/dev/null"],
        capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    assert result.returncode == 0
    assert "Total entries:" in result.stdout
    assert "Series 2000:" in result.stdout
