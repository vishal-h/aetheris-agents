"""Tests for scripts/order_formatter.py — t3."""
import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

from order_formatter import (
    COL_AMOUNT,
    COL_COLOR,
    COL_ITEM,
    COL_LINE,
    COL_PRICE,
    COL_QTY,
    COL_SPECIAL,
    FEE_PLACEHOLDERS,
    FIRST_DATA_ROW,
    write_order_form,
)

USE_CASE_ROOT = Path(__file__).parent.parent
SAMPLES_DIR = USE_CASE_ROOT / "data" / "samples"
TEMPLATE_FILE = SAMPLES_DIR / "Updated_Boxy_MSRP_Sales_Order_Form.xlsx"


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def multi_sheet_template(tmp_path: Path) -> Path:
    """Template with multiple sheets — verifies sheet stripping."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2000 Order Form"
    ws.cell(11, COL_ITEM).value = "*ITEM"
    wb.create_sheet("1000 Price List")
    wb.create_sheet("2000 Price List")
    wb.create_sheet("3000 Order Form")
    path = tmp_path / "multi_template.xlsx"
    wb.save(path)
    return path


@pytest.fixture()
def minimal_template(tmp_path: Path) -> Path:
    """Minimal workbook with a '2000 Order Form' sheet for unit tests."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2000 Order Form"
    # Row 11: headers (only B, C, D, E, F, K matter for the formatter)
    ws.cell(11, COL_LINE).value = "Line"
    ws.cell(11, COL_ITEM).value = "*ITEM"
    ws.cell(11, COL_COLOR).value = "*Color"
    ws.cell(11, COL_QTY).value = "*QTY"
    ws.cell(11, COL_PRICE).value = "Unit Price"
    ws.cell(11, COL_AMOUNT).value = "Amount"
    ws.cell(11, COL_SPECIAL).value = "Modification or Special Request Details"
    path = tmp_path / "template.xlsx"
    wb.save(path)
    return path


def _resolved_item(
    code: str = "W2739",
    drawing: str = "El1",
    qty: int = 1,
    unit_price: float = 450.8,
    color_code: str = "2001",
    color_name: str = "Ivory White",
    confidence: str = "exact",
    match_notes: str | None = None,
) -> dict:
    return {
        "component": {"code": code, "drawing": drawing, "qty": qty, "notes": None},
        "catalog_item": {
            "sku": f"{code}-{color_code}",
            "series": "2000",
            "color_code": color_code,
            "color_name": color_name,
            "description": "Wall Cabinet",
            "cabinet_type": "Wall Cabinet",
            "width_in": 27.0,
            "height_in": 39.0,
            "depth_in": 12.0,
            "msrp": unit_price,
        },
        "qty": qty,
        "unit_price": unit_price,
        "line_total": unit_price * qty,
        "match_confidence": confidence,
        "match_notes": match_notes,
    }


def _unresolved_item(code: str = "BLB42FHL", drawing: str = "El1", qty: int = 1) -> dict:
    return {
        "component": {"code": code, "drawing": drawing, "qty": qty, "notes": None},
        "catalog_item": None,
        "qty": qty,
        "unit_price": 0.0,
        "line_total": 0.0,
        "match_confidence": "unresolved",
        "match_notes": None,
    }


# ---------------------------------------------------------------------------
# Sheet stripping
# ---------------------------------------------------------------------------


def test_output_has_single_sheet(multi_sheet_template, tmp_path):
    write_order_form([], multi_sheet_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    assert wb.sheetnames == ["2000 Order Form"]


# ---------------------------------------------------------------------------
# Line item writing
# ---------------------------------------------------------------------------


def test_item_code_written_to_item_column(minimal_template, tmp_path):
    write_order_form([_resolved_item(code="W2739")], minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    assert ws.cell(FIRST_DATA_ROW, COL_ITEM).value == "W2739"


def test_color_code_written(minimal_template, tmp_path):
    write_order_form([_resolved_item(color_code="2001")], minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    assert ws.cell(FIRST_DATA_ROW, COL_COLOR).value == "2001"


def test_qty_written(minimal_template, tmp_path):
    write_order_form([_resolved_item(qty=3)], minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    assert ws.cell(FIRST_DATA_ROW, COL_QTY).value == 3


def test_unit_price_written(minimal_template, tmp_path):
    write_order_form([_resolved_item(unit_price=450.8)], minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    assert ws.cell(FIRST_DATA_ROW, COL_PRICE).value == pytest.approx(450.8)


def test_line_numbers_sequential(minimal_template, tmp_path):
    items = [_resolved_item(code=c) for c in ("W2739", "SB42", "DB30")]
    write_order_form(items, minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    for i, expected in enumerate([1, 2, 3], start=FIRST_DATA_ROW):
        assert ws.cell(i, COL_LINE).value == expected


# ---------------------------------------------------------------------------
# Amount formula
# ---------------------------------------------------------------------------


def test_amount_column_is_formula(minimal_template, tmp_path):
    write_order_form([_resolved_item()], minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    amount_val = ws.cell(FIRST_DATA_ROW, COL_AMOUNT).value
    assert isinstance(amount_val, str) and amount_val.startswith("="), (
        f"Amount must be a formula, got {amount_val!r}"
    )


def test_amount_formula_references_correct_row(minimal_template, tmp_path):
    items = [_resolved_item(), _resolved_item(code="SB42")]
    write_order_form(items, minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    f1 = ws.cell(FIRST_DATA_ROW, COL_AMOUNT).value
    f2 = ws.cell(FIRST_DATA_ROW + 1, COL_AMOUNT).value
    assert str(FIRST_DATA_ROW) in f1
    assert str(FIRST_DATA_ROW + 1) in f2


# ---------------------------------------------------------------------------
# Unresolved items
# ---------------------------------------------------------------------------


def test_unresolved_item_code_written(minimal_template, tmp_path):
    write_order_form([_unresolved_item("BLB42FHL")], minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    assert ws.cell(FIRST_DATA_ROW, COL_ITEM).value == "BLB42FHL"


def test_unresolved_item_price_blank(minimal_template, tmp_path):
    write_order_form([_unresolved_item()], minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    assert ws.cell(FIRST_DATA_ROW, COL_PRICE).value is None


def test_unresolved_item_color_blank(minimal_template, tmp_path):
    write_order_form([_unresolved_item()], minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    assert ws.cell(FIRST_DATA_ROW, COL_COLOR).value is None


def test_unresolved_item_special_request_flag(minimal_template, tmp_path):
    write_order_form([_unresolved_item()], minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    special = ws.cell(FIRST_DATA_ROW, COL_SPECIAL).value
    assert special is not None
    assert "UNRESOLVED - manual review required" in special


def test_match_notes_included_in_special_request(minimal_template, tmp_path):
    item = _resolved_item(match_notes="matched as W2439 after suffix strip", confidence="fuzzy")
    write_order_form([item], minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    special = ws.cell(FIRST_DATA_ROW, COL_SPECIAL).value
    assert special is not None
    assert "matched as W2439 after suffix strip" in special


def test_resolved_item_no_special_request_if_no_notes(minimal_template, tmp_path):
    write_order_form([_resolved_item(match_notes=None)], minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    assert ws.cell(FIRST_DATA_ROW, COL_SPECIAL).value is None


# ---------------------------------------------------------------------------
# Fee placeholder rows
# ---------------------------------------------------------------------------


def test_fee_rows_follow_line_items(minimal_template, tmp_path):
    items = [_resolved_item(code="W2739")]
    write_order_form(items, minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    fee_row_start = FIRST_DATA_ROW + len(items)
    fee_items_found = [
        ws.cell(fee_row_start + i, COL_ITEM).value for i in range(3)
    ]
    assert fee_items_found == list(FEE_PLACEHOLDERS)


def test_fee_rows_have_blank_price(minimal_template, tmp_path):
    write_order_form([_resolved_item()], minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    fee_start = FIRST_DATA_ROW + 1
    for i in range(3):
        assert ws.cell(fee_start + i, COL_PRICE).value is None
        assert ws.cell(fee_start + i, COL_AMOUNT).value is None


def test_three_fee_placeholders_present(minimal_template, tmp_path):
    write_order_form([], minimal_template, "test", tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "test_order_form.xlsx")
    ws = wb["2000 Order Form"]
    fee_items = [ws.cell(FIRST_DATA_ROW + i, COL_ITEM).value for i in range(3)]
    assert fee_items == ["Assembly Fee", "Modification Fee", "Delivery Fee"]


# ---------------------------------------------------------------------------
# Output file
# ---------------------------------------------------------------------------


def test_output_file_created(minimal_template, tmp_path):
    out = write_order_form([_resolved_item()], minimal_template, "MyProject", tmp_path)
    assert out.name == "MyProject_order_form.xlsx"
    assert out.exists()


def test_output_dir_created_if_missing(minimal_template, tmp_path):
    out_dir = tmp_path / "new_subdir"
    assert not out_dir.exists()
    write_order_form([], minimal_template, "test", out_dir)
    assert out_dir.exists()


# ---------------------------------------------------------------------------
# Integration tests — require sample files
# ---------------------------------------------------------------------------


@pytest.mark.integration
def test_cli_produces_output_file(tmp_path):
    elevation_pdf = SAMPLES_DIR / "Joey-_Kitchen_2D_Plans_V2.pdf"
    floor_pdf = SAMPLES_DIR / "Joey-_Kitchen_Plan_V2.pdf"

    t1 = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "plan_extractor.py"),
         str(elevation_pdf), str(floor_pdf)],
        capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    assert t1.returncode == 0

    t2 = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "catalog_resolver.py"),
         "--catalog", str(TEMPLATE_FILE),
         "--upper-finish", "2001:Ivory White:2000",
         "--lower-finish", "2004:Mingo Oak:2000"],
        input=t1.stdout, capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    assert t2.returncode == 0

    t3 = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "order_formatter.py"),
         "--template", str(TEMPLATE_FILE),
         "--project", "Joey_Kitchen_V2",
         "--output-dir", str(tmp_path)],
        input=t2.stdout, capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    assert t3.returncode == 0, f"t3 failed:\n{t3.stderr}"

    out_file = tmp_path / "Joey_Kitchen_V2_order_form.xlsx"
    assert out_file.exists(), f"Expected output file not found: {out_file}"


@pytest.mark.integration
def test_cli_line_item_count(tmp_path):
    elevation_pdf = SAMPLES_DIR / "Joey-_Kitchen_2D_Plans_V2.pdf"
    floor_pdf = SAMPLES_DIR / "Joey-_Kitchen_Plan_V2.pdf"

    t1 = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "plan_extractor.py"),
         str(elevation_pdf), str(floor_pdf)],
        capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    t2 = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "catalog_resolver.py"),
         "--catalog", str(TEMPLATE_FILE),
         "--upper-finish", "2001:Ivory White:2000",
         "--lower-finish", "2004:Mingo Oak:2000"],
        input=t1.stdout, capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    t3 = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "order_formatter.py"),
         "--template", str(TEMPLATE_FILE),
         "--project", "Joey_Kitchen_V2",
         "--output-dir", str(tmp_path)],
        input=t2.stdout, capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    assert t3.returncode == 0

    wb = openpyxl.load_workbook(tmp_path / "Joey_Kitchen_V2_order_form.xlsx")
    ws = wb["2000 Order Form"]

    # Count populated ITEM cells starting from row 12, excluding fee placeholders
    fee_names = set(FEE_PLACEHOLDERS)
    line_items = [
        ws.cell(r, COL_ITEM).value
        for r in range(FIRST_DATA_ROW, FIRST_DATA_ROW + 60)
        if ws.cell(r, COL_ITEM).value and ws.cell(r, COL_ITEM).value not in fee_names
    ]
    assert len(line_items) >= 10, f"Expected ≥10 line items, got {len(line_items)}"


@pytest.mark.integration
def test_cli_amount_column_has_formulas(tmp_path):
    elevation_pdf = SAMPLES_DIR / "Joey-_Kitchen_2D_Plans_V2.pdf"
    floor_pdf = SAMPLES_DIR / "Joey-_Kitchen_Plan_V2.pdf"

    t1 = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "plan_extractor.py"),
         str(elevation_pdf), str(floor_pdf)],
        capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    t2 = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "catalog_resolver.py"),
         "--catalog", str(TEMPLATE_FILE),
         "--upper-finish", "2001:Ivory White:2000",
         "--lower-finish", "2004:Mingo Oak:2000"],
        input=t1.stdout, capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    t3 = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "order_formatter.py"),
         "--template", str(TEMPLATE_FILE),
         "--project", "Joey_Kitchen_V2",
         "--output-dir", str(tmp_path)],
        input=t2.stdout, capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    assert t3.returncode == 0

    wb = openpyxl.load_workbook(tmp_path / "Joey_Kitchen_V2_order_form.xlsx")
    ws = wb["2000 Order Form"]

    # All Amount cells for populated rows must be formulas
    formula_count = 0
    for r in range(FIRST_DATA_ROW, FIRST_DATA_ROW + 60):
        item_val = ws.cell(r, COL_ITEM).value
        if item_val and item_val not in FEE_PLACEHOLDERS:
            amount_val = ws.cell(r, COL_AMOUNT).value
            assert isinstance(amount_val, str) and amount_val.startswith("="), (
                f"Row {r} Amount is not a formula: {amount_val!r}"
            )
            formula_count += 1
    assert formula_count >= 1


@pytest.mark.integration
def test_cli_unresolved_items_flagged(tmp_path):
    elevation_pdf = SAMPLES_DIR / "Joey-_Kitchen_2D_Plans_V2.pdf"
    floor_pdf = SAMPLES_DIR / "Joey-_Kitchen_Plan_V2.pdf"

    t1 = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "plan_extractor.py"),
         str(elevation_pdf), str(floor_pdf)],
        capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    t2 = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "catalog_resolver.py"),
         "--catalog", str(TEMPLATE_FILE),
         "--upper-finish", "2001:Ivory White:2000",
         "--lower-finish", "2004:Mingo Oak:2000"],
        input=t1.stdout, capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    t3 = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "order_formatter.py"),
         "--template", str(TEMPLATE_FILE),
         "--project", "Joey_Kitchen_V2",
         "--output-dir", str(tmp_path)],
        input=t2.stdout, capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    assert t3.returncode == 0

    wb = openpyxl.load_workbook(tmp_path / "Joey_Kitchen_V2_order_form.xlsx")
    ws = wb["2000 Order Form"]

    unresolved_flagged = [
        r for r in range(FIRST_DATA_ROW, FIRST_DATA_ROW + 60)
        if ws.cell(r, COL_SPECIAL).value
        and "UNRESOLVED - manual review required" in str(ws.cell(r, COL_SPECIAL).value)
    ]
    assert len(unresolved_flagged) >= 1, "Expected at least one unresolved item flagged"


@pytest.mark.integration
def test_no_stale_formulas_beyond_used_rows(tmp_path):
    """Rows 42–67 must be blank after formatter clears the unused template range."""
    elevation_pdf = SAMPLES_DIR / "Joey-_Kitchen_2D_Plans_V2.pdf"
    floor_pdf = SAMPLES_DIR / "Joey-_Kitchen_Plan_V2.pdf"

    t1 = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "plan_extractor.py"),
         str(elevation_pdf), str(floor_pdf)],
        capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    t2 = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "catalog_resolver.py"),
         "--catalog", str(TEMPLATE_FILE),
         "--upper-finish", "2001:Ivory White:2000",
         "--lower-finish", "2004:Mingo Oak:2000"],
        input=t1.stdout, capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    t3 = subprocess.run(
        [sys.executable, str(USE_CASE_ROOT / "scripts" / "order_formatter.py"),
         "--template", str(TEMPLATE_FILE),
         "--project", "Joey_Kitchen_V2",
         "--output-dir", str(tmp_path)],
        input=t2.stdout, capture_output=True, text=True, cwd=str(USE_CASE_ROOT),
    )
    assert t3.returncode == 0

    wb = openpyxl.load_workbook(tmp_path / "Joey_Kitchen_V2_order_form.xlsx")
    ws = wb["2000 Order Form"]

    # Joey kitchen: 27 items + 3 fees → rows 12–41 used; rows 42–67 must be blank
    stale = [
        (r, c, ws.cell(r, c).value)
        for r in range(42, 68)
        for c in range(COL_ITEM, COL_SPECIAL + 1)
        if ws.cell(r, c).value is not None
    ]
    assert not stale, f"Stale values in unused rows: {stale[:5]}"
