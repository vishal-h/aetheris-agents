import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


def _thin_top():
    return Border(top=Side(style="thin"))


def _write_cell(ws_cell, cell_spec, col_type=None):
    value = cell_spec["value"]
    if col_type in ("currency", "number") and isinstance(value, str):
        try:
            value = float(value)
        except (ValueError, TypeError):
            pass
    ws_cell.value = value
    ws_cell.font = Font(bold=cell_spec["bold"])
    ws_cell.alignment = Alignment(horizontal=cell_spec["align"])
    if col_type == "currency":
        ws_cell.number_format = '"$"#,##0.00'
    elif col_type == "number":
        # No forced decimals: a count like 1 shows "1", not "1.00".
        ws_cell.number_format = "#,##0.##"


def generate_xlsx(doc_spec, output_path, base_file=None):
    """Render the doc spec to xlsx.

    When ``base_file`` is given, the workbook is opened from it and the renderer
    writes data into the existing sheets, leaving everything above ``header_row``
    and left of ``data_col_start`` untouched (the base file owns those — logo,
    headers/footers, branding). When ``base_file`` is None, a fresh workbook is
    created (m1 behaviour, unchanged).
    """
    base_mode = base_file is not None
    if base_mode:
        wb = load_workbook(base_file)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    data_col_start = doc_spec.get("data_col_start", 1)

    for sheet in doc_spec["sheets"]:
        name = sheet["name"]
        if base_mode:
            ws = wb[name] if name in wb.sheetnames else wb.create_sheet(title=name)
        else:
            ws = wb.create_sheet(title=name)

        columns = sheet["columns"]
        header_row = sheet["header_row"]

        # Column widths: only for renderer-owned columns (at/right of data_col_start).
        for i, col in enumerate(columns):
            ws.column_dimensions[get_column_letter(data_col_start + i)].width = col["width"]

        for mr in sheet.get("merge_ranges", []):
            # In base-file mode the rows above header_row belong to the base file
            # (branding); skip merge ranges that fall there so we don't overwrite
            # the logo/company-name region.
            if base_mode and mr["row"] < header_row:
                continue
            ws.merge_cells(
                start_row=mr["row"], start_column=mr["col_start"],
                end_row=mr["row"], end_column=mr["col_end"],
            )
            top_left = ws.cell(row=mr["row"], column=mr["col_start"])
            top_left.value = mr["value"]
            top_left.font = Font(bold=True)
            top_left.alignment = Alignment(horizontal="center")

        for row_offset, row in enumerate(sheet["rows"]):
            physical_row = header_row + row_offset
            is_aggregate = row["type"] == "aggregate"

            for col_offset, cell_spec in enumerate(row["cells"]):
                col_idx = data_col_start + col_offset
                col_type = columns[col_offset]["type"] if col_offset < len(columns) else None
                ws_cell = ws.cell(row=physical_row, column=col_idx)
                _write_cell(ws_cell, cell_spec, col_type)
                if is_aggregate:
                    ws_cell.border = _thin_top()

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", default="output")
    parser.add_argument("--filename", default="document")
    parser.add_argument("--input", default=None, help="doc spec JSON file path (default: stdin)")
    parser.add_argument("--base-file", default=None,
                        help="xlsx base file to open and write into (branding preserved); "
                             "default: create a fresh workbook")
    args = parser.parse_args()

    try:
        src = open(args.input, encoding="utf-8") if args.input else sys.stdin
        doc_spec = json.load(src)
    except Exception as e:
        print(json.dumps({"status": "error", "error": str(e)}), file=sys.stderr)
        sys.exit(1)

    try:
        output_dir = Path(args.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{args.filename}.xlsx"
        generate_xlsx(doc_spec, output_path, base_file=args.base_file)
        print(str(output_path))
    except Exception as e:
        print(json.dumps({"status": "error", "error": str(e)}), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
